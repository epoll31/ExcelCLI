import click
import openpyxl
import pick

@click.command()
@click.argument('file', type=click.File('r'))

def main(file):
    click.echo('Welcome to the authentication system. Press any key to continue.')
    click.getchar()
    active_user = None
    while True:
        wb = openpyxl.load_workbook(file.name)
        ws = wb.active
        options = ['Log In', 'Sign Up', 'Change File', 'Exit']
        if active_user is not None:
            options = ['Log Out', 'Remove User', 'Change File', 'Exit']

        val, _ = pick.pick(options, 'Choose an option:', '>')
        match val:
            case 'Log In':
                click.echo('Log In')
                pairs = [(ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value) for i in range(2, ws.max_row + 1)]
                while active_user is None:
                    username = click.prompt('Username')
                    password = click.prompt('Password', hide_input=True)
                    if (username, password) in pairs:
                        active_user = (username, password)
                        click.echo('Logged in. Press any key to continue.')
                        click.getchar()
                    else:
                        click.echo('Invalid credentials')
                        # require y or n
                        #['y', 'Y', 'Yes', 'YES', 'n', 'N', 'no', 'NO'], 'Do you want to try again? (y/n)'
                        if click.confirm('Do you want to try again?', abort=False):
                            continue
                        else:
                            break

            case 'Sign Up':
                pairs = [(ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value) for i in range(2, ws.max_row + 1)]
                username = click.prompt('Username')
                password = click.prompt('Password', hide_input=True)
                if (username, password) in pairs:
                    click.echo('User already exists. Press any key to continue.')
                    click.getchar()
                else:
                    ws.cell(row=ws.max_row + 1, column=1, value=username)
                    ws.cell(row=ws.max_row, column=2, value=password)
                    wb.save(file.name)
                    click.echo('User created. Press any key to continue.')
                    click.getchar()
            case 'Log Out':
                active_user = None
                click.echo('Logged out. Press any key to continue.')
                click.getchar()
            case 'Remove User':
                if click.confirm('Are you sure you want to remove your account?', abort=False):
                    ws.delete_rows([i for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=1).value == active_user[0]][0])
                    active_user = None
                    wb.save(file.name)
                    click.echo('User removed. Press any key to continue.')
                else:
                    click.echo('User not removed. Press any key to continue.')
                click.getchar()
            case 'Change File':
                file = click.prompt('Enter the file name:', type=click.File('r'))
            case  'Exit':
                exit()

if __name__ == '__main__':
    main()